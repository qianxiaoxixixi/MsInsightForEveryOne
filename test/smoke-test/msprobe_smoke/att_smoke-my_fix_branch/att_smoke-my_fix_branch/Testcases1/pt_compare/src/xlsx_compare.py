import openpyxl
import sys

"""
Compare two Excel files (.xlsx) to check if they are identical.
:param file1: Path to the first Excel file.
:param file2: Path to the second Excel file.
:return: True if the files are identical, False otherwise.
"""
file1 = sys.argv[1]
file2 = sys.argv[2]
try:
    wb1 = openpyxl.load_workbook(file1)
    wb2 = openpyxl.load_workbook(file2)

    # Compare sheet names
    sheets1 = wb1.sheetnames
    sheets2 = wb2.sheetnames

    if sheets1 != sheets2:
        print("Sheet names do not match.")
        print(f"File1 sheets: {sheets1}")
        print(f"File2 sheets: {sheets2}")

    # Compare each sheet's content
    for sheet_name in sheets1:
        sheet1 = wb1[sheet_name]
        sheet2 = wb2[sheet_name]

        # Compare dimensions
        if sheet1.max_row != sheet2.max_row or sheet1.max_column != sheet2.max_column:
            print(f"Mismatch in sheet dimensions for sheet '{sheet_name}'.")

        # Compare cell values
        for row in range(1, sheet1.max_row + 1):
            for col in range(1, sheet1.max_column + 1):
                value1 = sheet1.cell(row=row, column=col).value
                value2 = sheet2.cell(row=row, column=col).value

                if value1 != value2:
                    print(f"Mismatch found in sheet '{sheet_name}' at cell ({row}, {col}):")
                    print(f"File1 value: {value1}")
                    print(f"File2 value: {value2}")

    print("The files are identical.")

except Exception as e:
    print(f"Error comparing files: {e}")
